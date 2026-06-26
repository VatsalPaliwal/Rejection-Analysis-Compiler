import streamlit as st
import pandas as pd
import re
from io import BytesIO
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
                
st.set_page_config(page_title="Rejection Analysis Compiler", layout="wide")
st.title("Rejection Analysis Compiler")
st.markdown("Upload the monthly workbook to generate a compiled rejection report.")

DEFECT_COLS = [
    "FPPA", "U/O SIZE", "SHRINKAGE", "B.SPOT", "FLASH", "COLOR",
    "CRACK", "SHORT", "BURN MARK", "EJECT. MARK", "WELD LINE",
    "BEND", "S.STREAK", "BOP ISSUE", "SCRATCH", "OTHER"
]

workbook = st.file_uploader("Upload Monthly Workbook", type=["xlsx", "xls"])


def find_header_row(df_raw):
    for i, row in df_raw.iterrows():
        vals = [str(v).strip().upper() for v in row.values]
        if "DATE" in vals and "COMPONENT" in vals:
            return i
    return 0


def read_sheet(xl, sheet_name):
    try:
        raw = xl.parse(sheet_name, header=None)
        header_row = find_header_row(raw)
        df = xl.parse(sheet_name, header=header_row)
        df.columns = [str(c).strip().upper() for c in df.columns]
        df = df.rename(columns={
            "TOTAL    REJECTION": "TOTAL REJECTION",
            "TOTAL  REJECTION": "TOTAL REJECTION"
        })
        df = df[df["COMPONENT"].notna() & (df["COMPONENT"].astype(str).str.strip() != "")]
        df["DATE"] = pd.to_datetime(df["DATE"], dayfirst=True, errors="coerce")
        df["COMPONENT"] = df["COMPONENT"].astype(str).str.strip().str.upper()
        df["CUSTOMER"] = df["CUSTOMER"].astype(str).str.strip().str.upper() if "CUSTOMER" in df.columns else ""
        df["TOTAL PRODUCTION"] = pd.to_numeric(df.get("TOTAL PRODUCTION", 0), errors="coerce").fillna(0)
        df["TOTAL REJECTION"] = pd.to_numeric(df.get("TOTAL REJECTION", 0), errors="coerce").fillna(0)
        for col in DEFECT_COLS:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
            else:
                df[col] = 0
        return df
    except Exception as e:
        st.error(f"Error reading sheet '{sheet_name}': {e}")
        return None


def nearest_future_match(process_df, final_df):
    process_df = process_df.reset_index(drop=True)
    final_df = final_df.reset_index(drop=True)
    final_df["_matched"] = False

    results = []

    for _, p_row in process_df.iterrows():
        comp = p_row["COMPONENT"]
        p_date = p_row["DATE"]

        candidates = final_df[
            (final_df["COMPONENT"] == comp) &
            (final_df["DATE"] >= p_date) &
            (~final_df["_matched"])
        ].copy()

        if not candidates.empty:
            candidates["_gap"] = (candidates["DATE"] - p_date).dt.days
            best_idx = candidates["_gap"].idxmin()
            f_row = final_df.loc[best_idx]
            final_df.at[best_idx, "_matched"] = True
            results.append(build_combined_row(p_row, f_row))
        else:
            results.append(build_process_only_row(p_row))

    unmatched_final = final_df[~final_df["_matched"]]
    for _, f_row in unmatched_final.iterrows():
        results.append(build_final_only_row(f_row))

    return pd.DataFrame(results)


def calc_rej(row):
    total_prod = row["TOTAL PRODUCTION"]
    total_rej = row["TOTAL REJECTION"]
    row["REJ. %"] = round((total_rej / total_prod) * 100, 2) if total_prod > 0 else 0
    row["REJ. PPM"] = round((total_rej / total_prod) * 1_000_000, 0) if total_prod > 0 else 0
    return row


def build_combined_row(p_row, f_row):
    row = {
        "DATE PROCESS": p_row["DATE"],
        "DATE FINAL": f_row["DATE"],
        "COMPONENT": p_row["COMPONENT"],
        "CUSTOMER": p_row.get("CUSTOMER") or f_row.get("CUSTOMER", ""),
        "TOTAL PRODUCTION": p_row["TOTAL PRODUCTION"] + f_row["TOTAL PRODUCTION"],
        "TOTAL REJECTION": p_row["TOTAL REJECTION"] + f_row["TOTAL REJECTION"],
    }
    for col in DEFECT_COLS:
        row[col] = p_row.get(col, 0) + f_row.get(col, 0)
    return calc_rej(row)


def build_process_only_row(p_row):
    row = {
        "DATE PROCESS": p_row["DATE"],
        "DATE FINAL": None,
        "COMPONENT": p_row["COMPONENT"],
        "CUSTOMER": p_row.get("CUSTOMER", ""),
        "TOTAL PRODUCTION": p_row["TOTAL PRODUCTION"],
        "TOTAL REJECTION": p_row["TOTAL REJECTION"],
    }
    for col in DEFECT_COLS:
        row[col] = p_row.get(col, 0)
    return calc_rej(row)


def build_final_only_row(f_row):
    row = {
        "DATE PROCESS": None,
        "DATE FINAL": f_row["DATE"],
        "COMPONENT": f_row["COMPONENT"],
        "CUSTOMER": f_row.get("CUSTOMER", ""),
        "TOTAL PRODUCTION": f_row["TOTAL PRODUCTION"],
        "TOTAL REJECTION": f_row["TOTAL REJECTION"],
    }
    for col in DEFECT_COLS:
        row[col] = f_row.get(col, 0)
    return calc_rej(row)


if st.button("Compile", type="primary"):
    if not workbook:
        st.warning("Please upload a workbook.")
    else:
        xl = pd.ExcelFile(workbook)
        sheet_names = xl.sheet_names

        process_sheet = next((s for s in sheet_names if re.search(r'process', s, re.IGNORECASE)), None)
        final_sheet = next((s for s in sheet_names if re.search(r'final', s, re.IGNORECASE)), None)

        if not process_sheet:
            st.error("No sheet matching 'process' found in workbook.")
        elif not final_sheet:
            st.error("No sheet matching 'final' found in workbook.")
        else:
            st.info(f"Process sheet: `{process_sheet}` | Final sheet: `{final_sheet}`")

            process_df = read_sheet(xl, process_sheet)
            final_df = read_sheet(xl, final_sheet)

            if process_df is not None and final_df is not None:
                with st.spinner("Compiling..."):
                    compiled = nearest_future_match(process_df, final_df)

                st.success(f"Compiled {len(compiled)} rows.")
                st.dataframe(compiled, use_container_width=True)

                # Write compiled sheet back into the original workbook
                output = BytesIO()
                workbook.seek(0)
                wb = openpyxl.load_workbook(workbook)

                # Remove existing COMPILED sheet if present
                if "COMPILED" in wb.sheetnames:
                    del wb["COMPILED"]

                # Write compiled df to a new sheet
                ws = wb.create_sheet("COMPILED")
                ws.append([])  # row 1 empty
                ws.append([])  # row 2 empty
                for r in dataframe_to_rows(compiled, index=False, header=True):
                    ws.append(r)
                for r in dataframe_to_rows(compiled, index=False, header=True):
                    ws.append(r)

                wb.save(output)
                output.seek(0)

                st.download_button(
                    label="Download Updated Workbook",
                    data=output,
                    file_name="compiled_rejection_analysis.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )